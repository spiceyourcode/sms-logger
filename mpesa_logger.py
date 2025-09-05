import re
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import os
import time
import json

class MPESATransactionLogger:
    def __init__(self, excel_file="mpesa_transactions.xlsx"):
        self.excel_file = excel_file
        self.setup_excel_file()
        
    def setup_excel_file(self):
        """Initialize the Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.excel_file):
            # Create new workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "M-PESA Transactions"
            
            # Define headers
            headers = [
                "Transaction Code", "Amount (KSh)", "Transaction Type", 
                "Recipient/Sender", "Date", "Time", "New Balance (KSh)",
                "Transaction Cost (KSh)", "Daily Limit Remaining (KSh)",
                "Raw Message", "Processed DateTime"
            ]
            
            # Add headers to first row
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            wb.save(self.excel_file)
            print(f"Created new Excel file: {self.excel_file}")
        else:
            print(f"Using existing Excel file: {self.excel_file}")
    
    def parse_mpesa_message(self, message):
        """Parse M-PESA transaction message and extract key information"""
        data = {}
        
        # Extract transaction code (usually at the beginning)
        code_match = re.search(r'^([A-Z0-9]{10})', message)
        data['transaction_code'] = code_match.group(1) if code_match else "N/A"
        
        # Extract amount
        amount_match = re.search(r'Ksh([\d,]+\.?\d*)', message)
        data['amount'] = amount_match.group(1).replace(',', '') if amount_match else "0"
        
        # Determine transaction type and recipient/sender
        if 'sent to' in message.lower():
            data['transaction_type'] = 'Send Money'
            recipient_match = re.search(r'sent to ([^on]+)', message, re.IGNORECASE)
            data['recipient_sender'] = recipient_match.group(1).strip() if recipient_match else "N/A"
        elif 'received from' in message.lower():
            data['transaction_type'] = 'Receive Money'
            sender_match = re.search(r'received from ([^on]+)', message, re.IGNORECASE)
            data['recipient_sender'] = sender_match.group(1).strip() if sender_match else "N/A"
        elif 'paid to' in message.lower():
            data['transaction_type'] = 'Pay Bill/Buy Goods'
            merchant_match = re.search(r'paid to ([^\.]+)', message, re.IGNORECASE)
            data['recipient_sender'] = merchant_match.group(1).strip() if merchant_match else "N/A"
        elif 'withdrawn' in message.lower():
            data['transaction_type'] = 'Withdraw'
            data['recipient_sender'] = 'ATM/Agent'
        else:
            data['transaction_type'] = 'Other'
            data['recipient_sender'] = 'N/A'
        
        # Extract date and time
        date_time_match = re.search(r'on (\d{1,2}/\d{1,2}/\d{2,4}) at (\d{1,2}:\d{2} [AP]M)', message)
        if date_time_match:
            data['date'] = date_time_match.group(1)
            data['time'] = date_time_match.group(2)
        else:
            data['date'] = "N/A"
            data['time'] = "N/A"
        
        # Extract new balance
        balance_match = re.search(r'New M-PESA balance is Ksh([\d,]+\.?\d*)', message)
        data['new_balance'] = balance_match.group(1).replace(',', '') if balance_match else "N/A"
        
        # Extract transaction cost
        cost_match = re.search(r'Transaction cost[,.]? Ksh([\d,]+\.?\d*)', message)
        data['transaction_cost'] = cost_match.group(1).replace(',', '') if cost_match else "0"
        
        # Extract daily limit remaining
        limit_match = re.search(r'Amount you can transact within the day is ([\d,]+\.?\d*)', message)
        data['daily_limit_remaining'] = limit_match.group(1).replace(',', '') if limit_match else "N/A"
        
        # Store raw message and processing time
        data['raw_message'] = message
        data['processed_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        return data
    
    def add_transaction_to_excel(self, transaction_data):
        """Add parsed transaction data to Excel file"""
        try:
            # Load existing workbook
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Find the next empty row
            next_row = ws.max_row + 1
            
            # Data to write (in order matching headers)
            row_data = [
                transaction_data['transaction_code'],
                transaction_data['amount'],
                transaction_data['transaction_type'],
                transaction_data['recipient_sender'],
                transaction_data['date'],
                transaction_data['time'],
                transaction_data['new_balance'],
                transaction_data['transaction_cost'],
                transaction_data['daily_limit_remaining'],
                transaction_data['raw_message'],
                transaction_data['processed_datetime']
            ]
            
            # Write data to Excel
            for col, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col, value=value)
            
            # Save the workbook
            wb.save(self.excel_file)
            print(f"✅ Transaction {transaction_data['transaction_code']} added to Excel successfully!")
            
        except Exception as e:
            print(f"❌ Error adding transaction to Excel: {str(e)}")
    
    def process_message(self, message):
        """Main method to process a single M-PESA message"""
        print(f"\n📱 Processing message: {message[:50]}...")
        
        # Parse the message
        transaction_data = self.parse_mpesa_message(message)
        
        # Add to Excel
        self.add_transaction_to_excel(transaction_data)
        
        # Print summary
        print(f"💰 Amount: KSh {transaction_data['amount']}")
        print(f"📅 Date/Time: {transaction_data['date']} {transaction_data['time']}")
        print(f"👤 Type: {transaction_data['transaction_type']}")
        print(f"🆔 Code: {transaction_data['transaction_code']}")
        
        return transaction_data

# Example usage and testing
if __name__ == "__main__":
    # Initialize the logger
    logger = MPESATransactionLogger("mpesa_transactions.xlsx")
    
    # Test message (your example)
    test_message = "THK04TF1W4 Confirmed. Ksh250.00 sent to Antony Kiumbe on 20/8/25 at 10:15 AM. New M-PESA balance is Ksh93.09. Transaction cost, Ksh7.00. Amount you can transact within the day is 499,700.00. Sign up for Lipa Na M-PESA Till online https://m-pesaforbusiness.co.ke"
    
    # Process the test message
    logger.process_message(test_message)
    
    print("\n" + "="*60)
    print("🚀 M-PESA Transaction Logger is ready!")
    print("📁 Excel file location:", os.path.abspath(logger.excel_file))
    print("="*60)
    
    # For real-time processing, you would integrate this with:
    # 1. SMS monitoring app/service
    # 2. WhatsApp Business API
    # 3. Email monitoring (if transactions come via email)
    # 4. Android app with SMS permissions
    
    # Example of batch processing multiple messages
    sample_messages = [
        "ABC123XYZ7 Confirmed. Ksh500.00 received from John Doe on 21/8/25 at 2:30 PM. New M-PESA balance is Ksh593.09. Transaction cost, Ksh0.00.",
        "DEF456GHI8 Confirmed. Ksh100.00 paid to KPLC PREPAID on 21/8/25 at 3:45 PM. New M-PESA balance is Ksh493.09. Transaction cost, Ksh0.00.",
    ]
    
    print("\n📊 Processing sample messages...")
    for msg in sample_messages:
        logger.process_message(msg)
        time.sleep(1)  # Small delay for demonstration